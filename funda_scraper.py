#!/usr/bin/env python3
"""
Funda Amsterdam Property Scraper v3 - Selenium
================================================
Funda uses bot detection that blocks plain HTTP requests.
This version uses Selenium with your real Chrome browser
to bypass the bot detection.

Usage:
    python funda_scraper.py

Requirements:
    pip install selenium webdriver-manager beautifulsoup4 lxml openpyxl geopy tqdm
    (Chrome browser must be installed)
"""

import json
import time
import logging
import re
import sys
from datetime import datetime
from typing import Optional, Dict, List, Tuple

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut, GeocoderServiceError
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    from tqdm import tqdm
except ImportError:
    def tqdm(it, **kw): return it

# ── Config ────────────────────────────────────────────────────────────────────

MAX_PAGES     = 20
PAGE_DELAY    = 3.0    # seconds between pages (be polite)
GEOCODE_DELAY = 1.0

def get_city_config():
    """Ask the user which city to scrape and return config."""
    print()
    print("=" * 55)
    print("  Funda Property Scraper")
    print("=" * 55)
    print()
    city = input("Which city do you want to scrape? (e.g. amsterdam, rotterdam, utrecht): ").strip().lower()
    if not city:
        city = "amsterdam"
        print(f"No city entered - defaulting to {city}")

    # Format city for Funda URL (lowercase, spaces to hyphens)
    city_slug = city.replace(" ", "-")
    city_label = city.title()

    search_url = (
        f"https://www.funda.nl/zoeken/koop/"
        f"?selected_area=%5B%22{city_slug}%22%5D&search_result={{page}}"
    )

    output_excel = f"funda_{city_slug}.xlsx"
    output_map   = f"funda_{city_slug}.html"
    log_file     = f"funda_{city_slug}.log"

    print(f"\nScraping: {city_label}")
    print(f"Output:   {output_excel} + {output_map}")
    print()

    return city_label, city_slug, search_url, output_excel, output_map, log_file

# ── Logging ───────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

# ── Driver ────────────────────────────────────────────────────────────────────

def make_driver() -> webdriver.Chrome:
    opts = Options()
    # Run visible (not headless) so Funda doesn't detect bot
    opts.add_argument("--window-size=1280,900")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    )
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=opts,
    )
    # Remove webdriver flag
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver

# ── Geocoder ──────────────────────────────────────────────────────────────────

geolocator = Nominatim(user_agent="funda-portfolio/1.0", timeout=10)
_geocache: Dict[str, Optional[Tuple[float, float]]] = {}

def geocode_postcode(postcode: str) -> Optional[Tuple[float, float]]:
    if not postcode:
        return None
    key = re.sub(r"\s+", "", postcode.upper())
    if key in _geocache:
        return _geocache[key]
    try:
        time.sleep(GEOCODE_DELAY)
        loc = geolocator.geocode(f"{key}, Netherlands", exactly_one=True, country_codes="nl")
        result = (round(loc.latitude, 6), round(loc.longitude, 6)) if loc else None
        _geocache[key] = result
        return result
    except (GeocoderTimedOut, GeocoderServiceError) as e:
        log.warning(f"Geocoder error for {postcode}: {e}")
        _geocache[key] = None
        return None

# ── Parsing ───────────────────────────────────────────────────────────────────

def clean(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())

def parse_price(text: str) -> Optional[int]:
    # Handles: € 575.000 k.k. / € 744.000 v.o.n. / € 1.200.000
    m = re.search(r"€\s*([\d\.]+)", text)
    if m:
        val = int(m.group(1).replace(".", ""))
        return val if val > 1000 else None  # sanity check
    return None

def parse_card(card) -> Optional[Dict]:
    """
    Parse one listing card.
    Confirmed structure from live inspection:
      div[class*="@container border-b"]
    Text format: Address / Postcode City / Price / m2 / Bedrooms / Energy / Agent
    """
    try:
        link = card.find("a", href=re.compile(r"/detail/koop/"))
        if not link:
            return None
        href = link.get("href", "")
        url  = "https://www.funda.nl" + href if href.startswith("/") else href

        lines = [l.strip() for l in card.get_text(separator="\n").split("\n") if l.strip()]

        if not lines:
            return None

        # Find address: must match Dutch street pattern
        # Valid:   "Keizersgracht 123"  "Van der Hoopstraat 12-A"  "Zuideinde 204"
        # Invalid: "Prachtige jaren 30 woning..."  "NVM Open Huizen Dag"
        # Rule: word(s) followed by a number at the END of the string (with optional suffix)
        address = ""
        for line in lines:
            # Must end with: number, or number-letter, or number/number (e.g. 12-A, 112-2, 3-hoog)
            if re.search(r"[A-Za-z]\s+\d+[-/\w]*$", line.strip()) and "€" not in line and "m" not in line[-3:]:
                address = line
                break

        if not address:
            return None

        postcode, city = "", "Amsterdam"
        for line in lines:
            m = re.match(r"^(\d{4}\s?[A-Z]{2})\s+(.+)$", line)
            if m:
                postcode = m.group(1).strip()
                city     = m.group(2).strip()
                break

        price = None
        for line in lines:
            # Catch both k.k. and v.o.n. price formats
            if "€" in line or ("k.k." in line.lower()) or ("v.o.n." in line.lower()):
                price = parse_price(line)
                if price:
                    break

        area_m2 = None
        for line in lines:
            m = re.match(r"^(\d+)\s*m", line)
            if m:
                area_m2 = int(m.group(1))
                break

        bedrooms = None
        found_area = False
        for line in lines:
            if re.match(r"^\d+\s*m", line):
                found_area = True
                continue
            if found_area and re.match(r"^\d+$", line):
                bedrooms = int(line)
                break

        energy = ""
        for line in lines:
            if re.match(r"^[A-G][+]{0,2}$", line):
                energy = line
                break

        agent = ""
        for line in reversed(lines):
            if line and not re.match(r"^[\d\s€m²+kk.,/%-]+$", line) and len(line) > 3:
                if line != address:
                    agent = line
                    break

        if not address and not price:
            return None

        return {
            "address":  clean(address),
            "postcode": postcode,
            "city":     city,
            "price":    price,
            "area_m2":  area_m2,
            "bedrooms": bedrooms,
            "energy":   energy,
            "agent":    clean(agent),
            "url":      url,
        }
    except Exception as e:
        log.debug(f"Card parse error: {e}")
        return None

def scrape_page(driver, page: int, search_url: str, city_label: str) -> List[Dict]:
    url = search_url.format(page=page)
    log.info(f"[{city_label}] Fetching page {page}...")
    driver.get(url)

    # Accept cookies if consent wall appears
    try:
        accept_btn = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH,
                "//button[contains(text(),'Alles accepteren') or contains(text(),'Accept all')]"
            ))
        )
        accept_btn.click()
        log.info("Cookie consent accepted")
        time.sleep(2)
    except Exception:
        pass  # No consent wall, continue

    # Wait for listings to load
    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'a[href*="/detail/koop/"]'))
        )
    except Exception:
        log.warning(f"Timeout waiting for listings on page {page}")

    time.sleep(2)  # Extra buffer for JS rendering

    soup  = BeautifulSoup(driver.page_source, "lxml")

    # All listings (regular + blikvanger) live as direct children of this wrapper div
    # Regular cards have class "@container border-b", blikvangers have no class
    # Using the wrapper as the source captures everything in one go
    all_cards = []
    seen_hrefs = set()

    for wrapper in soup.find_all("div", class_=lambda c: c and all(x in c for x in ["flex", "flex-col", "gap-3", "mt-4"])):
        for child in wrapper.find_all("div", recursive=False):
            link = child.find("a", href=re.compile(r"/detail/koop/"))
            if not link:
                continue
            href = link.get("href", "")
            if href in seen_hrefs:
                continue
            # Skip pure ad divs (no price)
            text = child.get_text()
            if "€" not in text and "v.o.n" not in text.lower():
                continue
            seen_hrefs.add(href)
            all_cards.append(child)

    # Fallback: if wrapper not found use regular cards
    if not all_cards:
        all_cards = soup.find_all("div", class_=lambda c: c and "@container border-b" in c)

    if not all_cards:
        log.warning(f"No cards on page {page}")
        with open(f"debug_selenium_page_{page}.html", "w", encoding="utf-8") as f:
            f.write(driver.page_source)
        return []

    results = []
    seen_on_page = set()
    for card in all_cards:
        listing = parse_card(card)
        if listing:
            # Deduplicate within the same page by URL
            if listing["url"] and listing["url"] in seen_on_page:
                continue
            seen_on_page.add(listing["url"])
            results.append(listing)
            log.info(
                f"  {listing['address']:38s} | "
                f"{'EUR '+str(listing['price']):>14s} | "
                f"{str(listing.get('area_m2','?'))+'m2':>7s} | "
                f"{listing['postcode']}"
            )

    log.info(f"Page {page}: {len(results)} listings")
    return results

# ── Excel ─────────────────────────────────────────────────────────────────────

def save_excel(listings: List[Dict], path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Amsterdam Listings"

    COLS = [
        ("address",  "Address"),
        ("postcode", "Postcode"),
        ("city",     "City"),
        ("price",    "Price (EUR)"),
        ("area_m2",  "Size (m2)"),
        ("bedrooms", "Bedrooms"),
        ("price_m2", "Price per m2"),
        ("energy",   "Energy Label"),
        ("agent",    "Agent"),
        ("lat",      "Latitude"),
        ("lon",      "Longitude"),
        ("geocoded", "Geocoded?"),
        ("url",      "Funda URL"),
    ]

    hf = PatternFill("solid", fgColor="E84E24")
    hn = Font(bold=True, color="FFFFFF", size=11)
    af = PatternFill("solid", fgColor="FFF0EC")

    for ci, (_, lbl) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=lbl)
        c.fill = hf; c.font = hn
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    for ri, l in enumerate(listings, 2):
        price = l.get("price")
        area  = l.get("area_m2")
        price_m2 = round(price / area) if price and area else None
        coords = l.get("_coords")
        lat = coords[0] if coords else None
        lon = coords[1] if coords else None
        vals = {
            "address":  l.get("address", ""),
            "postcode": l.get("postcode", ""),
            "city":     l.get("city", ""),
            "price":    price,
            "area_m2":  area,
            "bedrooms": l.get("bedrooms"),
            "price_m2": price_m2,
            "energy":   l.get("energy", ""),
            "agent":    l.get("agent", ""),
            "lat":      lat,
            "lon":      lon,
            "geocoded": "Yes" if coords else "No",
            "url":      l.get("url", ""),
        }
        fill = af if ri % 2 == 0 else None
        for ci, (key, _) in enumerate(COLS, 1):
            cell = ws.cell(row=ri, column=ci, value=vals[key])
            if fill: cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    widths = [35, 12, 15, 15, 12, 12, 15, 14, 28, 12, 12, 11, 55]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Summary")
    geocoded = sum(1 for l in listings if l.get("_coords"))
    prices   = [l["price"] for l in listings if l.get("price")]
    areas    = [l["area_m2"] for l in listings if l.get("area_m2")]
    ws2["A1"] = "Funda Amsterdam - Property Listings"
    ws2["A1"].font = Font(bold=True, size=13)
    for row, (lbl, val) in enumerate([
        ("Total listings", len(listings)),
        ("Geocoded",       geocoded),
        ("Avg price",      f"EUR {sum(prices)//len(prices):,}" if prices else "N/A"),
        ("Min price",      f"EUR {min(prices):,}" if prices else "N/A"),
        ("Max price",      f"EUR {max(prices):,}" if prices else "N/A"),
        ("Avg size",       f"{sum(areas)//len(areas)} m2" if areas else "N/A"),
        ("Scraped on",     datetime.now().strftime("%Y-%m-%d %H:%M")),
    ], start=3):
        ws2.cell(row=row, column=1, value=lbl).font = Font(bold=True)
        ws2.cell(row=row, column=2, value=val)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 35
    wb.save(path)
    log.info(f"Excel saved: {path}")

# ── Map ───────────────────────────────────────────────────────────────────────

def save_map(listings: List[Dict], path: str, city_label: str = "Amsterdam"):
    geocoded = [l for l in listings if l.get("_coords")]
    markers  = []
    for l in geocoded:
        lat, lon = l["_coords"]
        price = f"EUR {l['price']:,}" if l.get("price") else "N/A"
        area  = f"{l['area_m2']} m2" if l.get("area_m2") else "N/A"
        beds  = str(l.get("bedrooms", "?"))
        pm2   = f"EUR {l['price']//l['area_m2']:,}/m2" if l.get("price") and l.get("area_m2") else ""
        popup = (
            f"<b>{l.get('address','')}</b><br>"
            f"<span style='font-size:17px;color:#E84E24;font-weight:bold'>{price}</span><br>"
            f"{area} &bull; {beds} bed &bull; {pm2}<br>"
            f"<small>{l.get('agent','')}</small><br>"
            f"<a href='{l.get('url','#')}' target='_blank' style='color:#E84E24'>View on Funda</a>"
        )
        markers.append({"lat": lat, "lon": lon, "popup": popup, "price": l.get("price", 0)})

    markers_json = json.dumps(markers)
    total     = len(listings)
    geo_count = len(geocoded)
    prices    = [l["price"] for l in listings if l.get("price")]
    avg_price = f"EUR {sum(prices)//len(prices):,}" if prices else "N/A"

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Funda Amsterdam - Property Map</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
  *{{margin:0;padding:0;box-sizing:border-box}}
  body{{font-family:-apple-system,sans-serif;background:#f5f5f5}}
  #header{{background:#E84E24;color:white;padding:16px 24px;display:flex;align-items:center;justify-content:space-between;box-shadow:0 2px 8px rgba(0,0,0,.2)}}
  #header h1{{font-size:20px;font-weight:700}}
  #stats{{display:flex;gap:24px;background:white;padding:10px 24px;border-bottom:1px solid #eee;font-size:13px;color:#555}}
  #stats b{{color:#E84E24;font-size:15px}}
  #map{{height:calc(100vh - 100px)}}
  .leaflet-popup-content{{font-size:13px;line-height:1.7;min-width:190px}}
</style>
</head>
<body>
<div id="header">
  <div><h1>{city_label} Property Map</h1><p>Scraped from Funda.nl - {datetime.now().strftime("%d %B %Y")}</p></div>
  <div style="font-size:12px;opacity:.85;text-align:right">Built by Nick Zwart<br>Python + Selenium + Leaflet.js</div>
</div>
<div id="stats">
  <span>Listings: <b>{total}</b></span>
  <span>Mapped: <b>{geo_count}</b></span>
  <span>Avg price: <b>{avg_price}</b></span>
  <span>Click any pin for details</span>
</div>
<div id="map"></div>
<script>
const map = L.map('map').setView([52.3676,4.9041],13);
L.tileLayer('https://{{s}}.basemaps.cartocdn.com/rastertiles/voyager/{{z}}/{{x}}/{{y}}{{r}}.png',{{attribution:'CartoDB | Funda.nl',maxZoom:19,subdomains:'abcd'}}).addTo(map);
const markers={markers_json};
function pinColor(p){{
  if(!p) return '#888';
  if(p<400000) return '#22c55e';
  if(p<600000) return '#f59e0b';
  if(p<900000) return '#E84E24';
  return '#7c3aed';
}}
markers.forEach(m=>{{
  const c=pinColor(m.price);
  const icon=L.divIcon({{className:'',html:`<div style="width:13px;height:13px;border-radius:50%;background:${{c}};border:2px solid white;box-shadow:0 2px 4px rgba(0,0,0,.3)"></div>`,iconSize:[13,13],iconAnchor:[6,6]}});
  L.marker([m.lat,m.lon],{{icon}}).bindPopup(m.popup,{{maxWidth:230}}).addTo(map);
}});
const leg=L.control({{position:'bottomright'}});
leg.onAdd=()=>{{
  const d=L.DomUtil.create('div');
  d.style.cssText='background:white;padding:12px;border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,.15);font-size:12px;line-height:2';
  d.innerHTML='<b>Price range</b><br><span style="color:#22c55e">&#9679;</span> Under EUR 400k<br><span style="color:#f59e0b">&#9679;</span> EUR 400k-600k<br><span style="color:#E84E24">&#9679;</span> EUR 600k-900k<br><span style="color:#7c3aed">&#9679;</span> Over EUR 900k';
  return d;
}};
leg.addTo(map);
</script>
</body>
</html>"""

    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    log.info(f"Map saved: {path}")

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    # Ask for city and build config
    city_label, city_slug, search_url, output_excel, output_map, log_file = get_city_config()

    # Set up logging with city-specific log file
    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    logging.getLogger().addHandler(file_handler)

    log.info("=" * 55)
    log.info(f"  Funda {city_label} Scraper v3 - Selenium")
    log.info("=" * 55)
    log.info("A Chrome window will open - this is normal!")

    driver = make_driver()
    all_listings = []

    seen_urls = set()

    try:
        for page in range(1, MAX_PAGES + 1):
            listings = scrape_page(driver, page, search_url, city_label)
            if not listings:
                log.info(f"No listings on page {page} - stopping.")
                break

            # Detect duplicate pages (Funda returns same results when past last page)
            new_urls = {l["url"] for l in listings}
            overlap  = new_urls & seen_urls
            if overlap and len(overlap) >= len(new_urls) // 2:
                log.info(f"Page {page} is a duplicate of previous results - stopping.")
                break

            seen_urls.update(new_urls)
            all_listings.extend(listings)
            log.info(f"Total so far: {len(all_listings)}")
            time.sleep(PAGE_DELAY)
    finally:
        driver.quit()

    log.info(f"\nScraping complete: {len(all_listings)} listings.")

    if not all_listings:
        log.error(f"No listings found. Check debug_selenium_page_1.html")
        return

    log.info("Geocoding by postcode...")
    for l in tqdm(all_listings, desc="Geocoding"):
        l["_coords"] = geocode_postcode(l.get("postcode", ""))

    geocoded = sum(1 for l in all_listings if l.get("_coords"))
    log.info(f"Geocoded {geocoded}/{len(all_listings)}.")

    save_excel(all_listings, output_excel)
    save_map(all_listings, output_map, city_label)

    log.info("\n" + "=" * 55)
    log.info("DONE")
    log.info(f"  Excel : {output_excel}")
    log.info(f"  Map   : {output_map}  <- open in browser!")

if __name__ == "__main__":
    main()
